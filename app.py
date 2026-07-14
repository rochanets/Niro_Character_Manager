import csv
import hashlib
import io
import json
import os
import shutil
import socket
import sqlite3
import tempfile
import threading
import unicodedata
import webbrowser
import zipfile
from datetime import datetime
from uuid import uuid4

import requests
from dotenv import load_dotenv
from flask import (Flask, Response, abort, got_request_exception, jsonify, redirect,
                   render_template, request, send_file, stream_with_context, url_for)

from db import BASE_DIR, DB_PATH, UPLOAD_DIR, delete_upload, get_db, init_db, purge_expired_archive

load_dotenv(os.path.join(BASE_DIR, ".env"))

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 512 * 1024 * 1024  # 512 MB por request (backups completos em zip)


def _on_request_exception(sender, exception, **extra):
    """Registra qualquer exceção não tratada de uma requisição no módulo Logs,
    sem alterar o comportamento normal de erro do Flask (via sinal, não errorhandler)."""
    try:
        conn = get_db()
        log_event(conn, "error", "exception", f"{request.method} {request.path}: {exception}")
        conn.commit()
        conn.close()
    except Exception:
        pass


got_request_exception.connect(_on_request_exception, app)

ALLOWED_EXTS = {".png", ".jpg", ".jpeg", ".webp", ".gif", ".avif"}

PARAM_TYPES = {
    "region":      {"table": "regions",      "fk": "region_id",      "has_image": False, "label": "Região"},
    "affiliation": {"table": "affiliations", "fk": "affiliation_id", "has_image": False, "label": "Afiliação"},
    "element":     {"table": "elements",     "fk": "element_id",     "has_image": True,  "label": "Elemento"},
    "weapon":      {"table": "weapons",      "fk": "weapon_id",      "has_image": True,  "label": "Arma"},
    "role":        {"table": "roles",        "fk": None,             "has_image": False,
                     "has_description": True, "label": "Role"},
}

BANNER_LIMITS = {
    "unitario": {5: 1, 4: 3},
    "duplo":    {5: 2, 4: 3},
    "especial": {5: 12, 4: None},  # 4★ ilimitado em banners especiais
}

TEXT_LIMITS = {
    "dom": 500, "normal_attack": 500, "skill1": 500, "skill2": 500,
    "ultimate": 500, "personality": 4000, "profession": 200, "lore": 4000,
}



# ---------------------------------------------------------------- helpers

LOG_RETENTION = 2000


def log_event(conn, level, action, message):
    """Registra um evento no módulo Logs. `level` é info/success/warning/error."""
    conn.execute("INSERT INTO logs (level, action, message) VALUES (?, ?, ?)", (level, action, message))
    conn.execute("DELETE FROM logs WHERE id <= (SELECT MAX(id) FROM logs) - ?", (LOG_RETENTION,))


def save_image(file_storage, subdir):
    filename = file_storage.filename or "image.png"
    ext = os.path.splitext(filename)[1].lower() or ".png"
    if ext not in ALLOWED_EXTS:
        raise ValueError(f"Formato de imagem não suportado: {ext}")
    name = uuid4().hex + ext
    dest_dir = os.path.join(UPLOAD_DIR, subdir)
    os.makedirs(dest_dir, exist_ok=True)
    file_storage.save(os.path.join(dest_dir, name))
    return f"uploads/{subdir}/{name}"


def compose_reaction_image(elem_a_rel, elem_b_rel, size=160):
    """Gera o símbolo de uma reação: metade esquerda do ícone do elemento A
    e metade direita do ícone do elemento B, unidos ao centro."""
    from PIL import Image

    def half(rel, side):
        path = os.path.join(BASE_DIR, "static", rel.replace("/", os.sep))
        with Image.open(path) as im:
            im = im.convert("RGBA")
            scale = max(size / im.width, size / im.height)
            nw, nh = max(1, round(im.width * scale)), max(1, round(im.height * scale))
            im = im.resize((nw, nh), Image.LANCZOS)
            left, top = (nw - size) // 2, (nh - size) // 2
            im = im.crop((left, top, left + size, top + size))
            half_w = size // 2
            box = (0, 0, half_w, size) if side == "L" else (half_w, 0, size, size)
            return im.crop(box)

    canvas = Image.new("RGBA", (size, size), (0, 0, 0, 0))
    canvas.paste(half(elem_a_rel, "L"), (0, 0))
    canvas.paste(half(elem_b_rel, "R"), (size // 2, 0))
    dest_dir = os.path.join(UPLOAD_DIR, "reactions")
    os.makedirs(dest_dir, exist_ok=True)
    name = uuid4().hex + ".png"
    canvas.save(os.path.join(dest_dir, name))
    return f"uploads/reactions/{name}"


def fetch_params(conn):
    out = {}
    for key, meta in PARAM_TYPES.items():
        cols = "id, name"
        if meta["has_image"]:
            cols += ", image"
        if meta.get("has_description"):
            cols += ", description"
        rows = conn.execute(f"SELECT {cols} FROM {meta['table']} ORDER BY name COLLATE NOCASE").fetchall()
        out[key] = [dict(r) for r in rows]
    return out


def fetch_role_names(conn):
    return [r["name"] for r in conn.execute("SELECT name FROM roles ORDER BY name COLLATE NOCASE").fetchall()]


def fetch_role_descriptions(conn):
    return {r["name"]: r["description"] for r in conn.execute("SELECT name, description FROM roles").fetchall()}


def character_to_dict(row):
    d = dict(row)
    for key in ("region", "affiliation", "element", "weapon"):
        d[key] = {"id": d.pop(f"{key}_id"), "name": d.pop(f"{key}_name", None)}
        if f"{key}_image" in d:
            d[key]["image"] = d.pop(f"{key}_image")
    return d


CHAR_SELECT = """
SELECT c.*,
       r.name  AS region_name,
       a.name  AS affiliation_name,
       e.name  AS element_name,  e.image AS element_image,
       w.name  AS weapon_name,   w.image AS weapon_image
FROM characters c
LEFT JOIN regions r      ON r.id = c.region_id
LEFT JOIN affiliations a ON a.id = c.affiliation_id
LEFT JOIN elements e     ON e.id = c.element_id
LEFT JOIN weapons w      ON w.id = c.weapon_id
"""


# ---------------------------------------------------------------- páginas

@app.route("/")
def index():
    return redirect(url_for("page_chars"))


@app.route("/chars")
def page_chars():
    return render_template("chars.html", active="chars")


@app.route("/chars/new")
def page_char_new():
    conn = get_db()
    params = fetch_params(conn)
    roles = fetch_role_names(conn)
    conn.close()
    return render_template("character_form.html", active="chars",
                           params=params, character=None, limits=TEXT_LIMITS,
                           roles=roles)


@app.route("/chars/<int:char_id>")
def page_char_detail(char_id):
    conn = get_db()
    row = conn.execute(CHAR_SELECT + " WHERE c.id = ?", (char_id,)).fetchone()
    team = conn.execute(
        """SELECT t.name FROM team_members tm
           JOIN teams t ON t.id = tm.team_id
           WHERE tm.character_id = ?""", (char_id,)).fetchone()
    conn.close()
    if not row:
        abort(404)
    return render_template("character_detail.html", active="chars",
                           c=character_to_dict(row), team_name=team["name"] if team else None)


@app.route("/chars/<int:char_id>/edit")
def page_char_edit(char_id):
    conn = get_db()
    row = conn.execute(CHAR_SELECT + " WHERE c.id = ?", (char_id,)).fetchone()
    params = fetch_params(conn)
    roles = fetch_role_names(conn)
    conn.close()
    if not row:
        abort(404)
    return render_template("character_form.html", active="chars",
                           params=params, character=character_to_dict(row),
                           limits=TEXT_LIMITS, roles=roles)


@app.route("/parametros")
def page_params():
    return render_template("params.html", active="params")


@app.route("/times")
def page_teams():
    return render_template("teams.html", active="teams")


@app.route("/reacoes")
def page_reactions():
    return render_template("reactions.html", active="reactions")


@app.route("/banners")
def page_banners():
    return render_template("banners.html", active="banners")


@app.route("/historico")
def page_history():
    return render_template("history.html", active="history")


@app.route("/arquivo")
def page_archive():
    purge_expired_archive()
    return render_template("archive.html", active="archive")


@app.route("/logs")
def page_logs():
    return render_template("logs.html", active="logs")


# ---------------------------------------------------------------- API: parâmetros

@app.route("/api/params")
def api_params_all():
    conn = get_db()
    out = fetch_params(conn)
    conn.close()
    return jsonify(out)


@app.route("/api/params/<ptype>", methods=["POST"])
def api_param_create(ptype):
    meta = PARAM_TYPES.get(ptype) or abort(404)
    name = (request.form.get("name") or "").strip()
    if not name:
        return jsonify(error="Informe o nome."), 400
    image = None
    if meta["has_image"]:
        file = request.files.get("image")
        if not file or not file.filename:
            return jsonify(error=f"{meta['label']} precisa de uma imagem."), 400
        image = save_image(file, meta["table"])
    description = (request.form.get("description") or "").strip() if meta.get("has_description") else None
    conn = get_db()
    try:
        if meta["has_image"]:
            cur = conn.execute(f"INSERT INTO {meta['table']} (name, image) VALUES (?, ?)", (name, image))
        elif meta.get("has_description"):
            cur = conn.execute(f"INSERT INTO {meta['table']} (name, description) VALUES (?, ?)",
                               (name, description))
        else:
            cur = conn.execute(f"INSERT INTO {meta['table']} (name) VALUES (?)", (name,))
        log_event(conn, "success", "parametro.criado", f"{meta['label']} \"{name}\" cadastrado(a).")
        conn.commit()
        return jsonify(id=cur.lastrowid, name=name, image=image, description=description), 201
    except Exception:
        return jsonify(error=f"Já existe {meta['label']} com esse nome."), 409
    finally:
        conn.close()


@app.route("/api/params/<ptype>/<int:item_id>", methods=["PUT"])
def api_param_update(ptype, item_id):
    meta = PARAM_TYPES.get(ptype) or abort(404)
    conn = get_db()
    row = conn.execute(f"SELECT * FROM {meta['table']} WHERE id = ?", (item_id,)).fetchone()
    if not row:
        conn.close()
        abort(404)
    name = (request.form.get("name") or row["name"]).strip()
    old_name = row["name"]
    try:
        conn.execute(f"UPDATE {meta['table']} SET name = ? WHERE id = ?", (name, item_id))
        if meta["has_image"]:
            file = request.files.get("image")
            if file and file.filename:
                delete_upload(row["image"])
                image = save_image(file, meta["table"])
                conn.execute(f"UPDATE {meta['table']} SET image = ? WHERE id = ?", (image, item_id))
        if meta.get("has_description"):
            description = (request.form.get("description") or "").strip()
            conn.execute(f"UPDATE {meta['table']} SET description = ? WHERE id = ?", (description, item_id))
        if ptype == "role" and name != old_name:
            conn.execute("UPDATE characters SET role1 = ? WHERE role1 = ?", (name, old_name))
            conn.execute("UPDATE characters SET role2 = ? WHERE role2 = ?", (name, old_name))
        log_event(conn, "info", "parametro.editado", f"{meta['label']} \"{old_name}\" atualizado(a).")
        conn.commit()
        return jsonify(ok=True)
    except Exception:
        return jsonify(error=f"Já existe {meta['label']} com esse nome."), 409
    finally:
        conn.close()


@app.route("/api/params/<ptype>/<int:item_id>", methods=["DELETE"])
def api_param_delete(ptype, item_id):
    meta = PARAM_TYPES.get(ptype) or abort(404)
    reassign_to = request.args.get("reassign_to", type=int)
    conn = get_db()
    row = conn.execute(f"SELECT * FROM {meta['table']} WHERE id = ?", (item_id,)).fetchone()
    if not row:
        conn.close()
        abort(404)

    if ptype == "role":
        name = row["name"]
        affected = conn.execute(
            "SELECT id, name, card_promo FROM characters "
            "WHERE (role1 = ? OR role2 = ?) AND archived = 0 ORDER BY name",
            (name, name),
        ).fetchall()
        if affected and reassign_to is None:
            conn.close()
            return jsonify(in_use=True, characters=[dict(a) for a in affected]), 409
        if reassign_to is not None:
            new_role = conn.execute("SELECT name FROM roles WHERE id = ?", (reassign_to,)).fetchone()
            if reassign_to == item_id or not new_role:
                conn.close()
                return jsonify(error="Opção de substituição inválida."), 400
            conn.execute("UPDATE characters SET role1 = ? WHERE role1 = ?", (new_role["name"], name))
            conn.execute("UPDATE characters SET role2 = ? WHERE role2 = ?", (new_role["name"], name))
        conn.execute("DELETE FROM roles WHERE id = ?", (item_id,))
        log_event(conn, "warning", "parametro.excluido", f"{meta['label']} \"{name}\" excluído(a).")
        conn.commit()
        conn.close()
        return jsonify(ok=True)

    affected = conn.execute(
        f"SELECT id, name, card_promo FROM characters WHERE {meta['fk']} = ? AND archived = 0 ORDER BY name",
        (item_id,),
    ).fetchall()
    if affected and reassign_to is None:
        conn.close()
        return jsonify(in_use=True, characters=[dict(a) for a in affected]), 409
    if reassign_to is not None:
        if reassign_to == item_id or not conn.execute(
                f"SELECT 1 FROM {meta['table']} WHERE id = ?", (reassign_to,)).fetchone():
            conn.close()
            return jsonify(error="Opção de substituição inválida."), 400
        conn.execute(f"UPDATE characters SET {meta['fk']} = ? WHERE {meta['fk']} = ?",
                     (reassign_to, item_id))
    if meta["has_image"]:
        delete_upload(row["image"])
    conn.execute(f"DELETE FROM {meta['table']} WHERE id = ?", (item_id,))
    log_event(conn, "warning", "parametro.excluido", f"{meta['label']} \"{row['name']}\" excluído(a).")
    conn.commit()
    conn.close()
    return jsonify(ok=True)


# ---------------------------------------------------------------- API: backup (exportar/importar tudo)

def _db_snapshot_bytes():
    """Copia consistente do banco (via API de backup do sqlite3), mesmo com o app em uso."""
    fd, tmp_path = tempfile.mkstemp(suffix=".db")
    os.close(fd)
    try:
        src = sqlite3.connect(DB_PATH)
        dst = sqlite3.connect(tmp_path)
        with dst:
            src.backup(dst)
        src.close()
        dst.close()
        with open(tmp_path, "rb") as f:
            return f.read()
    finally:
        os.remove(tmp_path)


@app.route("/api/export")
def api_export():
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("niro.db", _db_snapshot_bytes())
        for root, _dirs, files in os.walk(UPLOAD_DIR):
            for name in files:
                full = os.path.join(root, name)
                arcname = os.path.join("uploads", os.path.relpath(full, UPLOAD_DIR))
                zf.write(full, arcname=arcname)
    buf.seek(0)
    filename = f"niro_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip"
    return send_file(buf, mimetype="application/zip", as_attachment=True, download_name=filename)


@app.route("/api/import", methods=["POST"])
def api_import():
    file = request.files.get("file")
    if not file or not file.filename:
        return jsonify(error="Envie um arquivo .zip."), 400
    try:
        zf = zipfile.ZipFile(file.stream)
    except zipfile.BadZipFile:
        return jsonify(error="Arquivo inválido: não é um .zip."), 400

    with tempfile.TemporaryDirectory() as tmp_dir:
        try:
            zf.extractall(tmp_dir)
        except Exception:
            return jsonify(error="Não foi possível extrair o .zip."), 400

        tmp_db = os.path.join(tmp_dir, "niro.db")
        if not os.path.isfile(tmp_db):
            return jsonify(error="ZIP inválido: niro.db não encontrado (use um backup gerado pelo próprio Niro)."), 400
        try:
            check = sqlite3.connect(tmp_db)
            check.execute("SELECT 1 FROM characters LIMIT 1")
            check.close()
        except sqlite3.Error:
            return jsonify(error="ZIP inválido: o banco de dados está corrompido ou não é do Niro."), 400

        shutil.copyfile(tmp_db, DB_PATH)

        tmp_uploads = os.path.join(tmp_dir, "uploads")
        if os.path.isdir(UPLOAD_DIR):
            shutil.rmtree(UPLOAD_DIR)
        if os.path.isdir(tmp_uploads):
            shutil.copytree(tmp_uploads, UPLOAD_DIR)
        else:
            os.makedirs(UPLOAD_DIR, exist_ok=True)

    init_db()
    conn = get_db()
    log_event(conn, "warning", "backup.importado", f"Backup \"{file.filename}\" importado (substituiu todos os dados).")
    conn.commit()
    conn.close()
    return jsonify(ok=True)


# ---------------------------------------------------------------- API: personagens

def parse_character_form(form, conn):
    data = {}
    for field in ("name", "age", "height", "dom", "normal_attack", "skill1",
                  "skill2", "ultimate", "personality", "profession", "lore"):
        value = (form.get(field) or "").strip()
        limit = TEXT_LIMITS.get(field)
        if limit and len(value) > limit:
            value = value[:limit]
        data[field] = value
    for field in ("region_id", "affiliation_id", "element_id", "weapon_id"):
        raw = form.get(field)
        data[field] = int(raw) if raw and raw.isdigit() else None
    valid_roles = set(fetch_role_names(conn))
    for field in ("role1", "role2"):
        value = (form.get(field) or "").strip()
        data[field] = value if value in valid_roles else ""
    data["rarity"] = int(form.get("rarity") or 0)
    return data


@app.route("/api/characters")
def api_characters():
    conn = get_db()
    rows = conn.execute(CHAR_SELECT + " WHERE c.archived = 0 ORDER BY c.name COLLATE NOCASE").fetchall()
    conn.close()
    return jsonify([character_to_dict(r) for r in rows])


@app.route("/api/characters", methods=["POST"])
def api_character_create():
    conn = get_db()
    data = parse_character_form(request.form, conn)
    if not data["name"]:
        conn.close()
        return jsonify(error="Informe o nome do personagem."), 400
    if data["rarity"] not in (4, 5):
        conn.close()
        return jsonify(error="Escolha a raridade."), 400
    card_full = request.files.get("card_full")
    card_promo = request.files.get("card_promo")
    if not card_full or not card_full.filename or not card_promo or not card_promo.filename:
        conn.close()
        return jsonify(error="Envie o card completo e o card promo."), 400
    if conn.execute("SELECT 1 FROM characters WHERE name = ? COLLATE NOCASE", (data["name"],)).fetchone():
        conn.close()
        return jsonify(error="Já existe um personagem com esse nome."), 409
    data["card_full"] = save_image(card_full, "characters")
    data["card_promo"] = save_image(card_promo, "characters")
    cols = ", ".join(data.keys())
    marks = ", ".join("?" for _ in data)
    cur = conn.execute(f"INSERT INTO characters ({cols}) VALUES ({marks})", list(data.values()))
    log_event(conn, "success", "personagem.criado", f"Personagem \"{data['name']}\" cadastrado.")
    conn.commit()
    conn.close()
    return jsonify(id=cur.lastrowid), 201


@app.route("/api/characters/<int:char_id>", methods=["PUT"])
def api_character_update(char_id):
    conn = get_db()
    row = conn.execute("SELECT * FROM characters WHERE id = ?", (char_id,)).fetchone()
    if not row:
        conn.close()
        abort(404)
    data = parse_character_form(request.form, conn)
    if not data["name"]:
        conn.close()
        return jsonify(error="Informe o nome do personagem."), 400
    if data["rarity"] not in (4, 5):
        conn.close()
        return jsonify(error="Escolha a raridade."), 400
    dup = conn.execute("SELECT 1 FROM characters WHERE name = ? COLLATE NOCASE AND id != ?",
                       (data["name"], char_id)).fetchone()
    if dup:
        conn.close()
        return jsonify(error="Já existe um personagem com esse nome."), 409
    for field, subkey in (("card_full", "card_full"), ("card_promo", "card_promo")):
        file = request.files.get(field)
        if file and file.filename:
            delete_upload(row[subkey])
            data[field] = save_image(file, "characters")
    sets = ", ".join(f"{k} = ?" for k in data)
    conn.execute(f"UPDATE characters SET {sets} WHERE id = ?", list(data.values()) + [char_id])
    log_event(conn, "info", "personagem.editado", f"Personagem \"{data['name']}\" atualizado.")
    conn.commit()
    conn.close()
    return jsonify(ok=True)


@app.route("/api/characters/<int:char_id>", methods=["DELETE"])
def api_character_archive(char_id):
    conn = get_db()
    row = conn.execute("SELECT id, name FROM characters WHERE id = ? AND archived = 0", (char_id,)).fetchone()
    if not row:
        conn.close()
        abort(404)
    conn.execute("UPDATE characters SET archived = 1, archived_at = datetime('now') WHERE id = ?", (char_id,))
    log_event(conn, "warning", "personagem.arquivado", f"Personagem \"{row['name']}\" movido para o Arquivo.")
    conn.commit()
    conn.close()
    return jsonify(ok=True)


@app.route("/api/characters/<int:char_id>/restore", methods=["POST"])
def api_character_restore(char_id):
    conn = get_db()
    row = conn.execute("SELECT name FROM characters WHERE id = ?", (char_id,)).fetchone()
    conn.execute("UPDATE characters SET archived = 0, archived_at = NULL WHERE id = ?", (char_id,))
    if row:
        log_event(conn, "success", "personagem.restaurado", f"Personagem \"{row['name']}\" restaurado do Arquivo.")
    conn.commit()
    conn.close()
    return jsonify(ok=True)


@app.route("/api/characters/<int:char_id>/permanent", methods=["DELETE"])
def api_character_permanent(char_id):
    conn = get_db()
    row = conn.execute("SELECT * FROM characters WHERE id = ? AND archived = 1", (char_id,)).fetchone()
    if not row:
        conn.close()
        return jsonify(error="Apenas personagens no arquivo podem ser excluídos definitivamente."), 400
    delete_upload(row["card_full"])
    delete_upload(row["card_promo"])
    conn.execute("DELETE FROM characters WHERE id = ?", (char_id,))
    log_event(conn, "warning", "personagem.excluido", f"Personagem \"{row['name']}\" excluído definitivamente.")
    conn.commit()
    conn.close()
    return jsonify(ok=True)


@app.route("/api/archive")
def api_archive():
    purge_expired_archive()
    conn = get_db()
    rows = conn.execute(CHAR_SELECT + " WHERE c.archived = 1 ORDER BY c.archived_at DESC").fetchall()
    conn.close()
    out = []
    for r in rows:
        d = character_to_dict(r)
        archived_at = datetime.strptime(d["archived_at"], "%Y-%m-%d %H:%M:%S")
        d["days_left"] = max(0, 30 - (datetime.utcnow() - archived_at).days)
        out.append(d)
    return jsonify(out)


# ---------------------------------------------------------------- API: importação de planilha

def _norm_header(value):
    text = unicodedata.normalize("NFKD", str(value or ""))
    text = text.encode("ascii", "ignore").decode()
    return " ".join(text.lower().split())


SHEET_FIELD_MAP = {
    "nome": "name", "nacao": "region", "regiao": "region", "elemento": "element",
    "raridade": "rarity", "arma": "weapon", "afiliacao": "affiliation",
    "idade": "age", "altura": "height", "ataque normal": "normal_attack",
    "skill 1": "skill1", "skill 2": "skill2", "ultimate": "ultimate",
    "personalidade": "personality", "lore": "lore", "dom": "dom",
    "profissao": "profession", "role 1": "role1", "role 2": "role2",
}


@app.route("/api/import_sheet", methods=["POST"])
def api_import_sheet():
    file = request.files.get("sheet")
    if not file or not file.filename:
        return jsonify(error="Envie um arquivo .xlsx ou .csv."), 400
    ext = os.path.splitext(file.filename)[1].lower()
    try:
        if ext == ".xlsx":
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(file.read()), read_only=True, data_only=True)
            table = [list(row) for row in wb.active.iter_rows(values_only=True)]
            wb.close()
        elif ext == ".csv":
            text = file.read().decode("utf-8-sig", errors="replace")
            sample = text[:2048]
            delim = ";" if sample.count(";") > sample.count(",") else ","
            table = list(csv.reader(io.StringIO(text), delimiter=delim))
        else:
            return jsonify(error="Formato não suportado. Envie .xlsx ou .csv."), 400
    except Exception as exc:
        return jsonify(error=f"Não foi possível ler a planilha ({exc})."), 400

    table = [row for row in table if row and any(str(c or "").strip() for c in row)]
    if len(table) < 2:
        return jsonify(error="A planilha precisa de um cabeçalho e ao menos uma linha de dados."), 400

    fields = [SHEET_FIELD_MAP.get(_norm_header(h)) for h in table[0]]
    if "name" not in fields:
        return jsonify(error="Planilha fora do padrão: a coluna NOME não foi encontrada."), 400

    rows = []
    for raw in table[1:]:
        item = {}
        for i, field in enumerate(fields):
            if not field or i >= len(raw):
                continue
            value = str(raw[i] if raw[i] is not None else "").strip()
            if not value:
                continue
            if field == "rarity":
                stars = value.count("★") or (int(value) if value.isdigit() else 0)
                if stars not in (4, 5):
                    continue
                value = str(stars)
            limit = TEXT_LIMITS.get(field)
            if limit and len(value) > limit:
                value = value[:limit]
            item[field] = value
        if item.get("name"):
            rows.append(item)
    if not rows:
        return jsonify(error="Nenhum personagem encontrado na planilha."), 400
    return jsonify(rows=rows)


# ---------------------------------------------------------------- thumbnails

THUMB_DIR = os.path.join(UPLOAD_DIR, ".thumbs")


@app.route("/thumb/<int:width>/<path:rel>")
def thumb_image(rel, width):
    """Serve uma versão reduzida (LANCZOS, alta qualidade) das imagens enviadas,
    evitando a distorção do downscale feito pelo navegador em imagens grandes."""
    width = max(16, min(width, 1600))
    src = os.path.normpath(os.path.join(BASE_DIR, "static", rel.replace("/", os.sep)))
    uploads_root = os.path.normpath(UPLOAD_DIR)
    if not src.startswith(uploads_root + os.sep) or not os.path.isfile(src):
        abort(404)
    if os.path.splitext(src)[1].lower() == ".gif":  # preserva animação
        return redirect(url_for("static", filename=rel))
    key = hashlib.sha1(f"{rel}|{width}".encode()).hexdigest() + ".webp"
    dest = os.path.join(THUMB_DIR, key)
    if not os.path.isfile(dest) or os.path.getmtime(dest) < os.path.getmtime(src):
        try:
            from PIL import Image
            with Image.open(src) as img:
                img = img.convert("RGBA")
                if img.width > width:
                    height = max(1, round(img.height * width / img.width))
                    img = img.resize((width, height), Image.LANCZOS)
                os.makedirs(THUMB_DIR, exist_ok=True)
                img.save(dest, "WEBP", quality=92, method=6)
        except Exception:
            return redirect(url_for("static", filename=rel))
    resp = send_file(dest, mimetype="image/webp", conditional=True)
    resp.headers["Cache-Control"] = "public, max-age=86400"
    return resp


# ---------------------------------------------------------------- API: banners

def banner_payload(conn):
    versions = {r["major"]: r["name"] for r in conn.execute("SELECT * FROM versions")}
    banners = []
    for b in conn.execute("SELECT * FROM banners ORDER BY major, minor, half"):
        chars = conn.execute(
            """SELECT c.id, c.name, c.rarity, c.card_promo
               FROM banner_characters bc JOIN characters c ON c.id = bc.character_id
               WHERE bc.banner_id = ? AND c.archived = 0
               ORDER BY c.rarity DESC, c.name COLLATE NOCASE""",
            (b["id"],),
        ).fetchall()
        banners.append({**dict(b), "characters": [dict(c) for c in chars]})
    return {"versions": versions, "banners": banners, "limits": BANNER_LIMITS}


@app.route("/api/banners")
def api_banners():
    conn = get_db()
    out = banner_payload(conn)
    conn.close()
    return jsonify(out)


@app.route("/api/banners", methods=["POST"])
def api_banner_create():
    body = request.get_json(force=True)
    major, minor = body.get("major"), body.get("minor")
    half = body.get("half", 1)
    btype = body.get("type")
    version_name = (body.get("version_name") or "").strip()
    if not isinstance(major, int) or not 1 <= major <= 8:
        return jsonify(error="Versão inválida."), 400
    if not isinstance(minor, int) or not 0 <= minor <= 8:
        return jsonify(error="Subversão inválida."), 400
    if btype not in BANNER_LIMITS:
        return jsonify(error="Tipo de banner inválido."), 400
    # banners especiais valem pela versão x.y inteira, não presos a uma metade,
    # e podem ser cadastrados mesmo com as 2 metades já ocupadas
    half = None if btype == "especial" else half
    if btype != "especial" and half not in (1, 2):
        return jsonify(error="Metade inválida."), 400
    conn = get_db()
    has_version = conn.execute("SELECT 1 FROM versions WHERE major = ?", (major,)).fetchone()
    if not has_version and not version_name:
        conn.close()
        return jsonify(error=f"A versão {major}.x ainda não tem nome. Informe um nome."), 400
    if half is not None and conn.execute(
            "SELECT 1 FROM banners WHERE major = ? AND minor = ? AND half = ?",
            (major, minor, half)).fetchone():
        conn.close()
        return jsonify(error=f"Já existe um banner na versão {major}.{minor} "
                             f"({'1ª' if half == 1 else '2ª'} metade)."), 409
    if not has_version:
        conn.execute("INSERT INTO versions (major, name) VALUES (?, ?)", (major, version_name))
    cur = conn.execute("INSERT INTO banners (major, minor, half, type) VALUES (?, ?, ?, ?)",
                       (major, minor, half, btype))
    log_event(conn, "success", "banner.criado", f"Banner ({btype}) criado na versão {major}.{minor}.")
    conn.commit()
    conn.close()
    return jsonify(id=cur.lastrowid), 201


def version_seq(major, minor):
    """Índice sequencial de versão (major.minor), crescente e contínuo entre majors."""
    return major * 9 + minor


@app.route("/api/banners/<int:banner_id>", methods=["PUT"])
def api_banner_update(banner_id):
    body = request.get_json(force=True)
    major, minor = body.get("major"), body.get("minor")
    half = body.get("half", 1)
    btype = body.get("type")
    version_name = (body.get("version_name") or "").strip()
    if not isinstance(major, int) or not 1 <= major <= 8:
        return jsonify(error="Versão inválida."), 400
    if not isinstance(minor, int) or not 0 <= minor <= 8:
        return jsonify(error="Subversão inválida."), 400
    if btype not in BANNER_LIMITS:
        return jsonify(error="Tipo de banner inválido."), 400
    # banners especiais valem pela versão x.y inteira, não presos a uma metade,
    # e podem ser cadastrados mesmo com as 2 metades já ocupadas
    half = None if btype == "especial" else half
    if btype != "especial" and half not in (1, 2):
        return jsonify(error="Metade inválida."), 400
    conn = get_db()
    banner = conn.execute("SELECT * FROM banners WHERE id = ?", (banner_id,)).fetchone()
    if not banner:
        conn.close()
        abort(404)
    has_version = conn.execute("SELECT 1 FROM versions WHERE major = ?", (major,)).fetchone()
    if not has_version and not version_name:
        conn.close()
        return jsonify(error=f"A versão {major}.x ainda não tem nome. Informe um nome."), 400
    if half is not None and conn.execute(
            "SELECT 1 FROM banners WHERE major = ? AND minor = ? AND half = ? AND id != ?",
            (major, minor, half, banner_id)).fetchone():
        conn.close()
        return jsonify(error=f"Já existe um banner na versão {major}.{minor} "
                             f"({'1ª' if half == 1 else '2ª'} metade)."), 409
    for rarity, limit in BANNER_LIMITS[btype].items():
        if limit is None:
            continue
        count = conn.execute(
            """SELECT COUNT(*) AS n FROM banner_characters bc JOIN characters c ON c.id = bc.character_id
               WHERE bc.banner_id = ? AND c.rarity = ?""", (banner_id, rarity)).fetchone()["n"]
        if count > limit:
            conn.close()
            return jsonify(error=f"O banner tem {count} personagens {rarity}★, mas o tipo escolhido "
                                 f"permite no máximo {limit}. Remova personagens antes de mudar o tipo."), 409
    # ao mudar de ciclo (major), garante que nenhum personagem do banner já apareça
    # em outro banner do ciclo de destino (um personagem = uma aparição por ciclo)
    if major != banner["major"]:
        char_ids = [r["character_id"] for r in conn.execute(
            "SELECT character_id FROM banner_characters WHERE banner_id = ?", (banner_id,)).fetchall()]
        for cid in char_ids:
            for a in conn.execute(
                    """SELECT b2.major, b2.minor FROM banner_characters bc
                       JOIN banners b2 ON b2.id = bc.banner_id
                       WHERE bc.character_id = ? AND bc.banner_id != ?""", (cid, banner_id)).fetchall():
                if a["major"] == major:
                    conn.close()
                    return jsonify(error=f"Mover para o ciclo {major}.x entraria em conflito "
                                         f"com personagens já usados na versão {a['major']}.{a['minor']}."), 409
    if not has_version:
        conn.execute("INSERT INTO versions (major, name) VALUES (?, ?)", (major, version_name))
    conn.execute("UPDATE banners SET major = ?, minor = ?, half = ?, type = ? WHERE id = ?",
                 (major, minor, half, btype, banner_id))
    log_event(conn, "info", "banner.editado", f"Banner #{banner_id} atualizado (versão {major}.{minor}).")
    conn.commit()
    conn.close()
    return jsonify(ok=True)


@app.route("/api/banners/<int:banner_id>", methods=["DELETE"])
def api_banner_delete(banner_id):
    conn = get_db()
    conn.execute("DELETE FROM banners WHERE id = ?", (banner_id,))
    log_event(conn, "warning", "banner.excluido", f"Banner #{banner_id} excluído.")
    conn.commit()
    conn.close()
    return jsonify(ok=True)


@app.route("/api/versions/<int:major>", methods=["PUT"])
def api_version_rename(major):
    name = (request.get_json(force=True).get("name") or "").strip()
    if not name:
        return jsonify(error="Informe o nome."), 400
    conn = get_db()
    conn.execute("INSERT INTO versions (major, name) VALUES (?, ?) "
                 "ON CONFLICT(major) DO UPDATE SET name = excluded.name", (major, name))
    conn.commit()
    conn.close()
    return jsonify(ok=True)


@app.route("/api/banners/<int:banner_id>/characters", methods=["POST"])
def api_banner_add_char(banner_id):
    char_id = request.get_json(force=True).get("character_id")
    conn = get_db()
    banner = conn.execute("SELECT * FROM banners WHERE id = ?", (banner_id,)).fetchone()
    char = conn.execute("SELECT * FROM characters WHERE id = ? AND archived = 0", (char_id,)).fetchone()
    if not banner or not char:
        conn.close()
        abort(404)
    if conn.execute("SELECT 1 FROM banner_characters WHERE banner_id = ? AND character_id = ?",
                    (banner_id, char_id)).fetchone():
        conn.close()
        return jsonify(error="Esse personagem já está no banner."), 409
    # um personagem só pode aparecer uma vez por ciclo (versão major x.*): se já
    # está em qualquer banner do mesmo ciclo (x.0, x.1, x.2, ...), fica bloqueado
    appearances = conn.execute(
        """SELECT b2.major, b2.minor FROM banner_characters bc
           JOIN banners b2 ON b2.id = bc.banner_id
           WHERE bc.character_id = ?""", (char_id,)).fetchall()
    for a in appearances:
        if a["major"] == banner["major"]:
            conn.close()
            return jsonify(error=f"Esse personagem já aparece no ciclo {banner['major']}.x "
                                 f"(versão {a['major']}.{a['minor']}) e só pode aparecer uma vez "
                                 f"por ciclo."), 409
    count = conn.execute(
        """SELECT COUNT(*) AS n FROM banner_characters bc
           JOIN characters c ON c.id = bc.character_id
           WHERE bc.banner_id = ? AND c.rarity = ?""",
        (banner_id, char["rarity"]),
    ).fetchone()["n"]
    limit = BANNER_LIMITS[banner["type"]][char["rarity"]]
    if limit is not None and count >= limit:
        conn.close()
        return jsonify(error=f"Limite de personagens {char['rarity']}★ atingido para banner "
                             f"{banner['type']} ({limit})."), 409
    conn.execute("INSERT INTO banner_characters (banner_id, character_id) VALUES (?, ?)",
                 (banner_id, char_id))
    conn.commit()
    conn.close()
    return jsonify(ok=True)


@app.route("/api/banners/<int:banner_id>/characters/<int:char_id>", methods=["DELETE"])
def api_banner_remove_char(banner_id, char_id):
    conn = get_db()
    conn.execute("DELETE FROM banner_characters WHERE banner_id = ? AND character_id = ?",
                 (banner_id, char_id))
    conn.commit()
    conn.close()
    return jsonify(ok=True)


# ---------------------------------------------------------------- API: times

TEAM_SIZE = 4
GRADIENT_MODES = 5


def team_payload(conn):
    reactions_all = [reaction_to_dict(r) for r in conn.execute(REACTION_SELECT)]
    reactions_by_pair = {tuple(sorted((r["element1"]["id"], r["element2"]["id"]))): r for r in reactions_all}
    reactions_by_id = {r["id"]: r for r in reactions_all}

    teams = []
    for t in conn.execute(
        """SELECT te.*, e1.name AS element1_name, e1.image AS element1_image,
                  e2.name AS element2_name, e2.image AS element2_image
           FROM teams te
           LEFT JOIN elements e1 ON e1.id = te.element1_id
           LEFT JOIN elements e2 ON e2.id = te.element2_id
           ORDER BY te.created_at, te.id"""
    ):
        rows = conn.execute(
            """SELECT tm.slot, c.id, c.name, c.rarity, c.card_promo, c.card_full, c.role1, c.role2,
                      e.name AS element_name, e.image AS element_image
               FROM team_members tm
               LEFT JOIN characters c ON c.id = tm.character_id
               LEFT JOIN elements e   ON e.id = c.element_id
               WHERE tm.team_id = ?""",
            (t["id"],),
        ).fetchall()
        by_slot = {r["slot"]: r for r in rows}
        members = []
        for slot in range(TEAM_SIZE):
            r = by_slot.get(slot)
            if r is None or r["id"] is None:
                members.append(None)
            else:
                members.append({
                    "id": r["id"], "name": r["name"], "rarity": r["rarity"],
                    "card_promo": r["card_promo"], "card_full": r["card_full"],
                    "role1": r["role1"], "role2": r["role2"],
                    "element_name": r["element_name"],
                    "element_image": r["element_image"],
                })
        d = dict(t)
        reaction_id = d.pop("reaction_id")
        reaction_cleared = d.pop("reaction_cleared")
        e1_id, e1_name, e1_img = d.pop("element1_id"), d.pop("element1_name"), d.pop("element1_image")
        e2_id, e2_name, e2_img = d.pop("element2_id"), d.pop("element2_name"), d.pop("element2_image")
        element1 = {"id": e1_id, "name": e1_name, "image": e1_img} if e1_id is not None else None
        element2 = {"id": e2_id, "name": e2_name, "image": e2_img} if e2_id is not None else None

        reaction = None
        if element1 and element2 and e1_id != e2_id:
            if reaction_id:
                reaction = reactions_by_id.get(reaction_id)
            elif not reaction_cleared:
                reaction = reactions_by_pair.get(tuple(sorted((e1_id, e2_id))))

        teams.append({**d, "element1": element1, "element2": element2, "members": members, "reaction": reaction})
    return teams


@app.route("/api/teams")
def api_teams():
    conn = get_db()
    out = team_payload(conn)
    conn.close()
    return jsonify(out)


def parse_element_ids(conn, body):
    """Valida element1_id (obrigatório) e element2_id (opcional, None = Random)."""
    e1 = body.get("element1_id")
    e2 = body.get("element2_id")
    if not isinstance(e1, int):
        return None, None, jsonify(error="Selecione o elemento 1 do time."), 400
    if e2 is not None and not isinstance(e2, int):
        return None, None, jsonify(error="Elemento 2 inválido."), 400
    ids = [e1] + ([e2] if e2 is not None else [])
    marks = ",".join("?" * len(ids))
    found = conn.execute(f"SELECT id FROM elements WHERE id IN ({marks})", ids).fetchall()
    if len(found) != len(set(ids)):
        return None, None, jsonify(error="Elemento inexistente."), 400
    return e1, e2, None, None


@app.route("/api/teams", methods=["POST"])
def api_team_create():
    body = request.get_json(force=True)
    name = (body.get("name") or "").strip()
    members = body.get("members")
    if not name:
        return jsonify(error="Informe o nome do time."), 400
    if not isinstance(members, list) or len(members) != TEAM_SIZE:
        return jsonify(error=f"O time precisa de exatamente {TEAM_SIZE} slots."), 400
    ids = [m for m in members if m is not None]
    if any(not isinstance(m, int) for m in ids):
        return jsonify(error="Personagem inválido."), 400
    if len(ids) != len(set(ids)):
        return jsonify(error="Um personagem não pode se repetir no mesmo time."), 400
    conn = get_db()
    e1, e2, err, code = parse_element_ids(conn, body)
    if err:
        conn.close()
        return err, code
    if conn.execute("SELECT 1 FROM teams WHERE name = ? COLLATE NOCASE", (name,)).fetchone():
        conn.close()
        return jsonify(error="Já existe um time com esse nome."), 409
    if ids:
        marks = ",".join("?" * len(ids))
        found = conn.execute(
            f"SELECT id FROM characters WHERE id IN ({marks}) AND archived = 0", ids).fetchall()
        if len(found) != len(ids):
            conn.close()
            return jsonify(error="Personagem inexistente ou arquivado."), 400
        clash = conn.execute(
            f"""SELECT c.name FROM team_members tm
                JOIN characters c ON c.id = tm.character_id
                WHERE tm.character_id IN ({marks})""", ids).fetchall()
        if clash:
            conn.close()
            names = ", ".join(r["name"] for r in clash)
            return jsonify(error=f"Personagem já está em outro time: {names}."), 409
    cur = conn.execute("INSERT INTO teams (name, element1_id, element2_id) VALUES (?, ?, ?)",
                       (name, e1, e2))
    team_id = cur.lastrowid
    for slot, member in enumerate(members):
        conn.execute("INSERT INTO team_members (team_id, slot, character_id) VALUES (?, ?, ?)",
                     (team_id, slot, member))
    log_event(conn, "success", "time.criado", f"Time \"{name}\" cadastrado.")
    conn.commit()
    conn.close()
    return jsonify(id=team_id), 201


@app.route("/api/teams/<int:team_id>", methods=["PUT"])
def api_team_update(team_id):
    body = request.get_json(force=True)
    name = (body.get("name") or "").strip()
    members = body.get("members")
    if not name:
        return jsonify(error="Informe o nome do time."), 400
    if not isinstance(members, list) or len(members) != TEAM_SIZE:
        return jsonify(error=f"O time precisa de exatamente {TEAM_SIZE} slots."), 400
    ids = [m for m in members if m is not None]
    if any(not isinstance(m, int) for m in ids):
        return jsonify(error="Personagem inválido."), 400
    if len(ids) != len(set(ids)):
        return jsonify(error="Um personagem não pode se repetir no mesmo time."), 400
    conn = get_db()
    if not conn.execute("SELECT 1 FROM teams WHERE id = ?", (team_id,)).fetchone():
        conn.close()
        abort(404)
    e1, e2, err, code = parse_element_ids(conn, body)
    if err:
        conn.close()
        return err, code
    if conn.execute("SELECT 1 FROM teams WHERE name = ? COLLATE NOCASE AND id != ?",
                     (name, team_id)).fetchone():
        conn.close()
        return jsonify(error="Já existe um time com esse nome."), 409
    if ids:
        marks = ",".join("?" * len(ids))
        found = conn.execute(
            f"SELECT id FROM characters WHERE id IN ({marks}) AND archived = 0", ids).fetchall()
        if len(found) != len(ids):
            conn.close()
            return jsonify(error="Personagem inexistente ou arquivado."), 400
        clash = conn.execute(
            f"""SELECT c.name FROM team_members tm
                JOIN characters c ON c.id = tm.character_id
                WHERE tm.character_id IN ({marks}) AND tm.team_id != ?""",
            ids + [team_id]).fetchall()
        if clash:
            conn.close()
            names = ", ".join(r["name"] for r in clash)
            return jsonify(error=f"Personagem já está em outro time: {names}."), 409
    conn.execute("UPDATE teams SET name = ?, element1_id = ?, element2_id = ? WHERE id = ?",
                 (name, e1, e2, team_id))
    conn.execute("DELETE FROM team_members WHERE team_id = ?", (team_id,))
    for slot, member in enumerate(members):
        conn.execute("INSERT INTO team_members (team_id, slot, character_id) VALUES (?, ?, ?)",
                     (team_id, slot, member))
    log_event(conn, "info", "time.editado", f"Time \"{name}\" atualizado.")
    conn.commit()
    conn.close()
    return jsonify(ok=True)


@app.route("/api/teams/<int:team_id>", methods=["DELETE"])
def api_team_delete(team_id):
    conn = get_db()
    row = conn.execute("SELECT name FROM teams WHERE id = ?", (team_id,)).fetchone()
    conn.execute("DELETE FROM teams WHERE id = ?", (team_id,))
    if row:
        log_event(conn, "warning", "time.excluido", f"Time \"{row['name']}\" excluído.")
    conn.commit()
    conn.close()
    return jsonify(ok=True)


@app.route("/api/teams/<int:team_id>/gradient", methods=["PUT"])
def api_team_gradient(team_id):
    mode = request.get_json(force=True).get("mode")
    if not isinstance(mode, int) or not 0 <= mode < GRADIENT_MODES:
        return jsonify(error="Modo de gradiente inválido."), 400
    conn = get_db()
    if not conn.execute("SELECT 1 FROM teams WHERE id = ?", (team_id,)).fetchone():
        conn.close()
        abort(404)
    conn.execute("UPDATE teams SET gradient_mode = ? WHERE id = ?", (mode, team_id))
    conn.commit()
    conn.close()
    return jsonify(ok=True)


@app.route("/api/teams/<int:team_id>/members/<int:char_id>", methods=["DELETE"])
def api_team_remove_member(team_id, char_id):
    conn = get_db()
    conn.execute("UPDATE team_members SET character_id = NULL "
                 "WHERE team_id = ? AND character_id = ?", (team_id, char_id))
    conn.commit()
    conn.close()
    return jsonify(ok=True)


@app.route("/api/teams/<int:team_id>/members/<int:slot>", methods=["PUT"])
def api_team_set_member(team_id, slot):
    if not 0 <= slot < TEAM_SIZE:
        abort(404)
    body = request.get_json(force=True)
    char_id = body.get("character_id")
    if not isinstance(char_id, int):
        return jsonify(error="Personagem inválido."), 400
    conn = get_db()
    if not conn.execute("SELECT 1 FROM teams WHERE id = ?", (team_id,)).fetchone():
        conn.close()
        abort(404)
    if not conn.execute("SELECT 1 FROM characters WHERE id = ? AND archived = 0",
                         (char_id,)).fetchone():
        conn.close()
        return jsonify(error="Personagem inexistente ou arquivado."), 400
    clash = conn.execute(
        """SELECT t.name FROM team_members tm JOIN teams t ON t.id = tm.team_id
           WHERE tm.character_id = ? AND tm.team_id != ?""", (char_id, team_id)).fetchone()
    if clash:
        conn.close()
        return jsonify(error=f'Personagem já está no time "{clash["name"]}".'), 409
    conn.execute("UPDATE team_members SET character_id = ? WHERE team_id = ? AND slot = ?",
                 (char_id, team_id, slot))
    conn.commit()
    conn.close()
    return jsonify(ok=True)


@app.route("/api/teams/<int:team_id>/reaction", methods=["PUT"])
def api_team_reaction_set(team_id):
    body = request.get_json(force=True)
    reaction_id = body.get("reaction_id")
    if not isinstance(reaction_id, int):
        return jsonify(error="Reação inválida."), 400
    conn = get_db()
    if not conn.execute("SELECT 1 FROM teams WHERE id = ?", (team_id,)).fetchone():
        conn.close()
        abort(404)
    if not conn.execute("SELECT 1 FROM reactions WHERE id = ?", (reaction_id,)).fetchone():
        conn.close()
        return jsonify(error="Reação inexistente."), 400
    conn.execute("UPDATE teams SET reaction_id = ?, reaction_cleared = 0 WHERE id = ?",
                 (reaction_id, team_id))
    conn.commit()
    conn.close()
    return jsonify(ok=True)


@app.route("/api/teams/<int:team_id>/reaction", methods=["DELETE"])
def api_team_reaction_clear(team_id):
    conn = get_db()
    if not conn.execute("SELECT 1 FROM teams WHERE id = ?", (team_id,)).fetchone():
        conn.close()
        abort(404)
    conn.execute("UPDATE teams SET reaction_id = NULL, reaction_cleared = 1 WHERE id = ?", (team_id,))
    conn.commit()
    conn.close()
    return jsonify(ok=True)


# ---------------------------------------------------------------- API: reações

REACTION_SELECT = """
    SELECT r.*, e1.name AS element1_name, e1.image AS element1_image,
           e2.name AS element2_name, e2.image AS element2_image
    FROM reactions r
    JOIN elements e1 ON e1.id = r.element1_id
    JOIN elements e2 ON e2.id = r.element2_id
"""


def reaction_to_dict(row):
    d = dict(row)
    d["element1"] = {"id": d.pop("element1_id"), "name": d.pop("element1_name"), "image": d.pop("element1_image")}
    d["element2"] = {"id": d.pop("element2_id"), "name": d.pop("element2_name"), "image": d.pop("element2_image")}
    return d


@app.route("/api/reactions")
def api_reactions():
    conn = get_db()
    rows = conn.execute(REACTION_SELECT + " ORDER BY r.id").fetchall()
    out = [reaction_to_dict(r) for r in rows]
    conn.close()
    return jsonify(out)


@app.route("/api/reactions", methods=["POST"])
def api_reaction_create():
    body = request.get_json(force=True)
    e1 = body.get("element1_id")
    e2 = body.get("element2_id")
    name = (body.get("name") or "").strip()
    description = (body.get("description") or "").strip()
    effect = (body.get("effect") or "").strip()
    if not isinstance(e1, int) or not isinstance(e2, int):
        return jsonify(error="Selecione os dois elementos."), 400
    if e1 == e2:
        return jsonify(error="Escolha dois elementos diferentes."), 400
    if not name:
        return jsonify(error="Informe o nome da reação."), 400
    lo, hi = sorted((e1, e2))
    conn = get_db()
    elems = conn.execute("SELECT id, image FROM elements WHERE id IN (?, ?)", (lo, hi)).fetchall()
    by_id = {r["id"]: r["image"] for r in elems}
    if len(by_id) != 2:
        conn.close()
        return jsonify(error="Elemento inexistente."), 400
    if not by_id.get(lo) or not by_id.get(hi):
        conn.close()
        return jsonify(error="Os dois elementos precisam ter uma imagem cadastrada em Parâmetros."), 400
    if conn.execute("SELECT 1 FROM reactions WHERE element1_id = ? AND element2_id = ?", (lo, hi)).fetchone():
        conn.close()
        return jsonify(error="Já existe uma reação cadastrada para esses elementos."), 409
    try:
        image = compose_reaction_image(by_id[lo], by_id[hi])
    except Exception:
        conn.close()
        return jsonify(error="Não foi possível gerar a imagem da reação."), 400
    cur = conn.execute(
        "INSERT INTO reactions (element1_id, element2_id, name, description, effect, image) "
        "VALUES (?, ?, ?, ?, ?, ?)", (lo, hi, name, description, effect, image))
    log_event(conn, "success", "reacao.criada", f"Reação \"{name}\" cadastrada.")
    conn.commit()
    reaction_id = cur.lastrowid
    conn.close()
    return jsonify(id=reaction_id), 201


@app.route("/api/reactions/<int:reaction_id>", methods=["PUT"])
def api_reaction_update(reaction_id):
    body = request.get_json(force=True)
    name = (body.get("name") or "").strip()
    description = (body.get("description") or "").strip()
    effect = (body.get("effect") or "").strip()
    if not name:
        return jsonify(error="Informe o nome da reação."), 400
    conn = get_db()
    if not conn.execute("SELECT 1 FROM reactions WHERE id = ?", (reaction_id,)).fetchone():
        conn.close()
        abort(404)
    conn.execute("UPDATE reactions SET name = ?, description = ?, effect = ? WHERE id = ?",
                 (name, description, effect, reaction_id))
    log_event(conn, "info", "reacao.editada", f"Reação \"{name}\" atualizada.")
    conn.commit()
    conn.close()
    return jsonify(ok=True)


@app.route("/api/reactions/<int:reaction_id>", methods=["DELETE"])
def api_reaction_delete(reaction_id):
    conn = get_db()
    row = conn.execute("SELECT image, name FROM reactions WHERE id = ?", (reaction_id,)).fetchone()
    if not row:
        conn.close()
        abort(404)
    conn.execute("DELETE FROM reactions WHERE id = ?", (reaction_id,))
    log_event(conn, "warning", "reacao.excluida", f"Reação \"{row['name']}\" excluída.")
    conn.commit()
    conn.close()
    delete_upload(row["image"])
    return jsonify(ok=True)


# ---------------------------------------------------------------- API: histórico

@app.route("/api/history")
def api_history():
    banner_id = request.args.get("banner_id", type=int)
    conn = get_db()
    banners = conn.execute(
        """SELECT b.id, b.major, b.minor, b.half,
                  (SELECT COUNT(*) FROM banner_characters bc
                   JOIN characters c ON c.id = bc.character_id
                   WHERE bc.banner_id = b.id AND c.archived = 0) AS n_chars
           FROM banners b ORDER BY b.major, b.minor, b.half"""
    ).fetchall()
    timeline = [b for b in banners if b["n_chars"] > 0]
    # dropdown lista do mais recente para o mais antigo (a timeline em si, usada no
    # cálculo de ausência, permanece em ordem crescente)
    options = [{"id": b["id"], "label": f"{b['major']}.{b['minor']} ({'1ª' if b['half'] == 1 else '2ª'})"}
               for b in reversed(timeline)]
    if banner_id is None or not any(b["id"] == banner_id for b in timeline):
        conn.close()
        return jsonify(options=options, rows=None)

    current_idx = next(i for i, b in enumerate(timeline) if b["id"] == banner_id)
    appearances = {}
    for i, b in enumerate(timeline[:current_idx + 1]):
        for row in conn.execute(
                """SELECT bc.character_id FROM banner_characters bc
                   JOIN characters c ON c.id = bc.character_id
                   WHERE bc.banner_id = ? AND c.archived = 0""", (b["id"],)):
            appearances[row["character_id"]] = i

    chars = conn.execute(
        "SELECT id, name, rarity, card_promo FROM characters WHERE archived = 0 "
        "ORDER BY rarity DESC, name COLLATE NOCASE").fetchall()
    conn.close()

    rows = []
    for c in chars:
        last = appearances.get(c["id"])
        if last is None:
            gap = current_idx + 1
            last_label = None
        else:
            gap = current_idx - last
            b = timeline[last]
            last_label = f"{b['major']}.{b['minor']} ({'1ª' if b['half'] == 1 else '2ª'})"
        rows.append({"id": c["id"], "name": c["name"], "rarity": c["rarity"],
                     "card_promo": c["card_promo"], "gap": gap, "last_banner": last_label})
    rows.sort(key=lambda r: (-r["rarity"], -r["gap"], r["name"].lower()))
    return jsonify(options=options, rows=rows,
                   current=next(o["label"] for o in options if o["id"] == banner_id))


# ---------------------------------------------------------------- API: IA (Google AI Studio / Gemini)

AI_FIELD_SPECS = {
    "normal_attack": ("Ataque Normal",
                      "Descreva o ataque normal (básico) do personagem em combate, coerente com a arma e o "
                      "elemento dele. Inclua um nome curto para o ataque e a descrição do golpe. "
                      "Máximo de 450 caracteres."),
    "skill1": ("Skill 1",
               "Crie a primeira habilidade ativa (Skill 1) do personagem, com nome e descrição do efeito em "
               "combate, coerente com elemento, arma e tema do personagem. Máximo de 450 caracteres."),
    "skill2": ("Skill 2",
               "Crie a segunda habilidade ativa (Skill 2) do personagem, com nome e descrição, diferente e "
               "complementar à Skill 1. Máximo de 450 caracteres."),
    "ultimate": ("Ultimate",
                 "Crie a habilidade suprema (Ultimate) do personagem: o golpe mais poderoso e espetacular, "
                 "com nome épico e descrição do efeito. Máximo de 450 caracteres."),
    "personality": ("Personalidade",
                    "Escreva a personalidade do personagem em profundidade: temperamento, maneirismos, "
                    "medos, valores, como trata aliados e inimigos, contradições internas. Texto corrido, "
                    "envolvente. Máximo de 3500 caracteres."),
    "lore": ("Lore",
             "Escreva a lore (história de fundo) do personagem: origem, eventos marcantes, conexão com sua "
             "região e afiliação, como obteve seu Dom e seu papel no mundo de Niro. Tom épico e misterioso, "
             "texto corrido. Máximo de 3500 caracteres."),
}

FIELD_LABELS_PT = {
    "name": "Nome", "age": "Idade", "height": "Altura", "rarity": "Raridade",
    "region": "Região", "affiliation": "Afiliação", "element": "Elemento", "weapon": "Arma",
    "role1": "Role 1 (função em combate)", "role2": "Role 2 (função secundária)",
    "dom": "Dom", "normal_attack": "Ataque Normal", "skill1": "Skill 1", "skill2": "Skill 2",
    "ultimate": "Ultimate", "personality": "Personalidade", "profession": "Profissão", "lore": "Lore",
}

COMBAT_FIELDS = {"normal_attack", "skill1", "skill2", "ultimate"}


@app.route("/api/ai_fill", methods=["POST"])
def api_ai_fill():
    api_key = (os.environ.get("GOOGLE_AI_API_KEY") or "").strip()
    if not api_key:
        return jsonify(error="Chave do Google AI Studio não configurada. Edite o arquivo .env e "
                             "preencha GOOGLE_AI_API_KEY."), 400
    body = request.get_json(force=True)
    field = body.get("field")
    if field not in AI_FIELD_SPECS:
        return jsonify(error="Campo inválido."), 400
    data = body.get("data") or {}
    draft = (str(body.get("draft") or "")).strip()

    label, instruction = AI_FIELD_SPECS[field]
    context_lines = []
    for key, pt in FIELD_LABELS_PT.items():
        value = (str(data.get(key) or "")).strip()
        if value and key != field:
            context_lines.append(f"- {pt}: {value}")
    context = "\n".join(context_lines) if context_lines else "(nenhum campo preenchido ainda)"

    extra = ""
    if field in COMBAT_FIELDS:
        conn = get_db()
        role_descriptions = fetch_role_descriptions(conn)
        conn.close()
        roles = [r for r in (str(data.get("role1") or "").strip(), str(data.get("role2") or "").strip())
                 if r in role_descriptions]
        if roles:
            role_lines = "\n".join(f"- {r}: {role_descriptions[r]}" for r in roles)
            extra += ("\n\nFunções (roles) do personagem em combate — a habilidade deve ser totalmente "
                      f"coerente com essas funções:\n{role_lines}\n"
                      "Lembre-se: neste jogo o ataque normal é ataque físico (não elemental), a menos que "
                      "alguma habilidade converta o ataque normal em dano elemental.")
    if draft:
        extra += ("\n\nO usuário escreveu o texto abaixo como base para este campo. Use-o como BASE "
                  "PRINCIPAL da resposta: preserve a intenção, os nomes e os detalhes descritos, apenas "
                  "estruturando, lapidando e completando o texto final, ainda coerente com as demais "
                  f"informações do personagem.\nBase do usuário: \"{draft}\"")

    system_prompt = (
        "Você é um escritor criativo de worldbuilding do projeto Niro, um universo de fantasia/RPG com "
        "regiões místicas, runas antigas, elementos mágicos e personagens jogáveis de 4 e 5 estrelas, no "
        "estilo de RPGs gacha. Você escreve em português do Brasil, com tom épico, criativo e coerente com "
        "as informações fornecidas. Responda APENAS com o texto solicitado para o campo, sem títulos, sem "
        "markdown, sem comentários adicionais e sem repetir o nome do campo."
    )
    user_prompt = (
        f"Informações já preenchidas do personagem:\n{context}\n\n"
        f"Tarefa: preencha o campo \"{label}\".\n{instruction}{extra}"
    )

    model = os.environ.get("GOOGLE_AI_MODEL", "gemini-flash-latest")
    payload = {
        "systemInstruction": {"parts": [{"text": system_prompt}]},
        "contents": [{"role": "user", "parts": [{"text": user_prompt}]}],
        "generationConfig": {"maxOutputTokens": 3000},
    }
    headers = {
        "X-goog-api-key": api_key,
        "Content-Type": "application/json",
    }
    url = (f"https://generativelanguage.googleapis.com/v1beta/models/"
           f"{model}:streamGenerateContent?alt=sse")

    def generate():
        try:
            with requests.post(url, json=payload, headers=headers,
                               stream=True, timeout=180) as resp:
                if resp.status_code != 200:
                    detail = resp.text[:300]
                    yield f"ERRO: Google AI Studio retornou {resp.status_code}. {detail}"
                    return
                # O stream SSE vem sem charset no Content-Type e o requests assume
                # ISO-8859-1, desconfigurando os acentos — força UTF-8.
                resp.encoding = "utf-8"
                for line in resp.iter_lines(decode_unicode=True):
                    if not line or not line.startswith("data: "):
                        continue
                    chunk = line[6:]
                    try:
                        delta = json.loads(chunk)["candidates"][0]["content"]["parts"][0].get("text") or ""
                    except (KeyError, IndexError, json.JSONDecodeError):
                        continue
                    if delta:
                        yield delta
        except requests.RequestException as exc:
            yield f"ERRO: falha de conexão com o Google AI Studio ({exc})"

    return Response(stream_with_context(generate()),
                    mimetype="text/plain; charset=utf-8",
                    headers={"X-Accel-Buffering": "no", "Cache-Control": "no-cache"})


# ---------------------------------------------------------------- API: logs

LOG_LEVELS = {"info", "success", "warning", "error"}


@app.route("/api/logs")
def api_logs():
    conn = get_db()
    level = request.args.get("level") or ""
    action_filter = (request.args.get("action") or "").strip()
    q = (request.args.get("q") or "").strip()
    limit = max(1, min(int(request.args.get("limit", 200)), 500))
    offset = max(0, int(request.args.get("offset", 0)))

    where, params = [], []
    if level in LOG_LEVELS:
        where.append("level = ?")
        params.append(level)
    if action_filter:
        where.append("action = ?")
        params.append(action_filter)
    if q:
        where.append("(message LIKE ? OR action LIKE ?)")
        params += [f"%{q}%", f"%{q}%"]
    clause = f"WHERE {' AND '.join(where)}" if where else ""

    total = conn.execute(f"SELECT COUNT(*) FROM logs {clause}", params).fetchone()[0]
    rows = conn.execute(
        f"SELECT * FROM logs {clause} ORDER BY id DESC LIMIT ? OFFSET ?",
        params + [limit, offset]).fetchall()
    actions = [r["action"] for r in conn.execute("SELECT DISTINCT action FROM logs ORDER BY action")]
    conn.close()
    return jsonify(items=[dict(r) for r in rows], total=total, actions=actions)


@app.route("/api/logs", methods=["DELETE"])
def api_logs_clear():
    conn = get_db()
    conn.execute("DELETE FROM logs")
    conn.commit()
    conn.close()
    return jsonify(ok=True)


# ---------------------------------------------------------------- bootstrap

def find_free_port(preferred=3004, tries=50):
    for port in range(preferred, preferred + tries):
        with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
            try:
                s.bind(("127.0.0.1", port))
                return port
            except OSError:
                continue
    return preferred


if __name__ == "__main__":
    init_db()
    purge_expired_archive()
    port = find_free_port(3004)
    threading.Timer(1.0, lambda: webbrowser.open(f"http://localhost:{port}")).start()
    print(f"\n  Niro Character Manager rodando em http://localhost:{port}\n")
    app.run(host="127.0.0.1", port=port, debug=False)
